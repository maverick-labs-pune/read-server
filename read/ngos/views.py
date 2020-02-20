#  Copyright (c) 2020. Maverick Labs
#    This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU Affero General Public License as,
#   published by the Free Software Foundation, either version 3 of the
#   License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#   along with this program.  If not, see <http://www.gnu.org/licenses/>.

import base64
# import the logging library
import logging

from datetime import datetime, timezone, timedelta
from django.contrib.auth.models import Group, Permission
from django.db import transaction, DatabaseError
from django.db.models import Q, F, CharField, Value
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from io import BytesIO
from openpyxl import load_workbook
from rest_framework import pagination
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.utils import json
from rest_framework.viewsets import ViewSet

from academic_years.models import AcademicYear
from books.models import Book, Inventory, BookLevel
from books.serializers import BookSerializer, InventorySerializer
from books.validators import validate_add_book_request, validate_book_file_excel_content, \
    validate_inventory_book_file_excel_content, validate_school_file_excel_content, \
    validate_book_and_inventory_file_excel_content
from classrooms.models import Classroom, ClassroomAcademicYear
from classrooms.serializers import ClassroomSerializer, ClassroomAcademicYearSerializer, ClassroomTableSerializer
from classrooms.validators import validate_add_classroom_request
from ngos.models import NGO, Level
from ngos.schemas import NGOSchema
from ngos.serializers import NGOSerializer, LevelSerializer
from ngos.utils import switch, student_sort_by_value, session_sort_by_value, user_sort_by_value
from ngos.validators import validate_change_ngo_request
from read.constants import GROUPS, GroupType, ERROR_403_JSON, FileType, USER_WORKSHEET_NAME, \
    BOOK_WORKSHEET_NAME, INVENTORY_BOOK_WORKSHEET_NAME, READ_SESSION_PENDING, READ_SESSION_EVALUATED_NOT_VERIFIED, \
    READ_SESSION_UPCOMING, SCHOOL_WORKSHEET_NAME, BOOK_LENDING, REGULAR, EVALUATION
from read.constants import SESSION_EVALUATED, SESSION_NON_EVALUATED
from read.utils import get_ngo_specific_group_name, write_to_excel, get_group_type_from_name, create_file_upload_error, \
    get_inventory_status, create_file_upload_serializer_error, get_group_type_from_request_user, \
    create_serializer_error, create_response_error, request_user_belongs_to_ngo, create_response_data, get_school_type, \
    get_school_medium, get_school_category, get_valid_school_categories, get_valid_school_types, \
    get_valid_school_mediums, get_book_level, get_valid_book_levels, get_valid_user_types, get_valid_inventory_statuses, \
    create_inventory
from read.validators import validate_deactivate_ngo_admin_request, validate_user_type
from read_sessions.models import ReadSession
from read_sessions.serializers import ReadSessionSerializer, ReadSessionClassroomSerializer, \
    ReadSessionBookFairySerializer, ReadSessionBookFairyReportSerializer
from read_sessions.validators import validate_add_session_request
from schools.models import SchoolCategory, SchoolType, SchoolMedium, School, Standard
from schools.serializers import SchoolSerializer
from schools.validators import validate_add_school_request
from users.models import User, SupervisorBookFairy
from users.permissions import has_permission, PERMISSION_CAN_VIEW_NGO, PERMISSION_CAN_ADD_NGO, PERMISSIONS_NGO_ADMIN, \
    PERMISSION_NGO_SUPERVISIOR, PERMISSIONS_NGO_BOOK_FAIRY, PERMISSIONS_READ_ADMIN, PERMISSION_CAN_CHANGE_NGO, \
    CanExportUsers, CanImportUsers, CanImportBook, CanExportBook, PERMISSION_CAN_DESTROY_NGO, CanViewNGO, CanChangeNGO, \
    CanAddBook, CanAddSchool, PERMISSION_CAN_CHANGE_LEVEL, PERMISSION_CAN_DESTROY_LEVEL, PERMISSION_CAN_ADD_LEVEL, \
    CanDeleteLevel, CanAddLevel, CanDeleteNGO, CanAddReadSession, PERMISSION_CAN_ADD_READ_SESSION_CLASSROOM, CanAddUser, \
    CanAddClassroom, CanExportSchool, CanImportSchool
from users.schemas import UserSchema
from users.serializers import UserSerializer, SupervisorBookFairySerializer
from users.validators import validate_user_file_excel_content, validate_user_file_import
from django.db.models.functions import Concat

# Get an instance of a logger
logger = logging.getLogger(__name__)


class ScheduledBookFairyException(Exception):
    def __init__(self, message):
        self.message = message


class NGOViewSet(ViewSet):

    def list(self, request):
        if not has_permission(request, PERMISSION_CAN_VIEW_NGO):
            return Response(status=403, data=ERROR_403_JSON())

        group_type = get_group_type_from_request_user(request.user)
        if group_type == GroupType.READ_ADMIN:
            queryset = NGO.objects.filter(is_active=True)
        else:
            queryset = NGO.objects.filter(is_active=True, user=request.user)

        serializer = NGOSerializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request):
        if not has_permission(request, PERMISSION_CAN_ADD_NGO):
            return Response(status=403, data=ERROR_403_JSON())

        serializer = NGOSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(status=400, data=create_serializer_error(serializer))

        try:
            with transaction.atomic():
                ngo = serializer.save()
                for group in GROUPS:
                    if group == GroupType.READ_ADMIN:
                        ngo_specific_group_name = group.value
                    else:
                        ngo_specific_group_name = group.value + " " + ngo.key
                    django_group, created = Group.objects.get_or_create(
                        name=ngo_specific_group_name,
                        defaults={'name': ngo_specific_group_name},
                    )

                    if group == GroupType.READ_ADMIN:
                        for code_name, name, _ in PERMISSIONS_READ_ADMIN:
                            permission = Permission.objects.get(codename=code_name, name=name)
                            django_group.permissions.add(permission)
                            logging.debug("Added  " + name + " permission to " + group.value)

                    elif group == GroupType.NGO_ADMIN:
                        for code_name, name, _ in PERMISSIONS_NGO_ADMIN:
                            permission = Permission.objects.get(codename=code_name, name=name)
                            django_group.permissions.add(permission)
                            logging.debug("Added  " + name + " permission to " + group.value)

                    elif group == GroupType.BOOK_FAIRY:
                        for code_name, name, _ in PERMISSIONS_NGO_BOOK_FAIRY:
                            permission = Permission.objects.get(codename=code_name, name=name)
                            django_group.permissions.add(permission)
                            logging.debug("Added  " + name + " permission to " + group.value)

                    elif group == GroupType.SUPERVISOR:
                        for code_name, name, _ in PERMISSION_NGO_SUPERVISIOR:
                            permission = Permission.objects.get(codename=code_name, name=name)
                            django_group.permissions.add(permission)
                            logging.debug("Added  " + name + " permission to " + group.value)

                # Create levels for ngo
                level_data = {"rank": 1, "en_in": "1", "mr_in": "1", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": True, "is_regular": True}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                level_data = {"rank": 2, "en_in": "2", "mr_in": "2", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": True, "is_regular": True}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                level_data = {"rank": 3, "en_in": "3", "mr_in": "3", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": True, "is_regular": True}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                level_data = {"rank": 4, "en_in": "4", "mr_in": "4", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": False, "is_regular": True}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                level_data = {"rank": 5, "en_in": "4.1", "mr_in": "4.1", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": True, "is_regular": False}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                level_data = {"rank": 6, "en_in": "4.2", "mr_in": "4.2", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": True, "is_regular": False}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                level_data = {"rank": 7, "en_in": "4.3", "mr_in": "4.3", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": True, "is_regular": False}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                level_data = {"rank": 8, "en_in": "4.4", "mr_in": "4.4", "ngo_id": ngo.id, "is_active": True,
                              "is_evaluation": True, "is_regular": False}
                serializer = LevelSerializer(data=level_data)
                if serializer.is_valid():
                    serializer.save()

                return Response(serializer.data, status=201)
        except (Permission.DoesNotExist, Group.DoesNotExist) as e:
            return Response(status=500, data=create_response_error(e))

    def retrieve(self, request, pk=None):
        if not has_permission(request, PERMISSION_CAN_VIEW_NGO):
            return Response(status=403, data=ERROR_403_JSON())
        queryset = NGO.objects.all()
        item = get_object_or_404(queryset, key=pk)
        serializer = NGOSerializer(item)
        return Response(serializer.data)

    def update(self, request, pk=None):
        if not has_permission(request, PERMISSION_CAN_CHANGE_NGO):
            return Response(status=403, data=ERROR_403_JSON())
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist:
            return Response(status=404)

        group_type = get_group_type_from_request_user(request.user)
        if group_type != GroupType.READ_ADMIN and request.user.ngo != ngo:
            return Response(status=403, data=ERROR_403_JSON())

        is_valid, error_message = validate_change_ngo_request(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        name = request.data.get('name')
        address = request.data.get('address')
        logo = request.data.get('logo')
        description = request.data.get('description')

        ngo.name = name.strip() if name else None
        ngo.address = address.strip() if address else None
        ngo.logo = logo.strip() if logo else None
        ngo.description = description.strip() if description else None
        ngo.save()
        serializer = NGOSerializer(ngo)
        return Response(serializer.data)

    @action(detail=True, methods=['POST'])
    def deactivate(self, request, pk=None):
        if not has_permission(request, PERMISSION_CAN_DESTROY_NGO):
            return Response(status=403, data=ERROR_403_JSON())

        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist:
            return Response(status=404)

        if get_group_type_from_request_user(request.user) == GroupType.READ_ADMIN:
            return Response(status=403, data=ERROR_403_JSON())
        ngo.is_active = False
        ngo.save()
        return Response(status=204)

    @action(detail=True, methods=['POST'], schema=NGOSchema.deactivate_ngo_admin(), permission_classes=[CanChangeNGO])
    def deactivate_admin(self, request, pk=None):
        # validate request
        is_valid, error_message = validate_deactivate_ngo_admin_request(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist:
            return Response(status=404)
        if not (request_user_belongs_to_ngo(request, ngo) or get_group_type_from_request_user(
                request.user) == GroupType.READ_ADMIN):
            return Response(status=403, data=ERROR_403_JSON())

        user_key = request.data['user_key']
        try:
            user = User.objects.get(key=user_key)
        except User.DoesNotExist:
            return Response(status=404)

        ngo_admin_group_name = get_ngo_specific_group_name(GroupType.NGO_ADMIN, ngo.key)
        admin_group = Group.objects.get(name=ngo_admin_group_name)
        user.groups.remove(admin_group)
        return Response(status=200)

    @action(detail=True, methods=['GET'], permission_classes=[CanViewNGO])
    def admins(self, request, pk=None):
        sort = request.GET.get('sort')
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist:
            return Response(status=404)
        paginator = pagination.PageNumberPagination()
        ngo_admin_group_name = get_ngo_specific_group_name(GroupType.NGO_ADMIN, ngo.key)
        admins = User.objects.filter(ngo__key=pk, groups__name=ngo_admin_group_name, is_active=True).order_by(sort)
        result = paginator.paginate_queryset(admins, request)
        serializer = UserSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['POST'], schema=NGOSchema.add_book(), permission_classes=[CanAddBook])
    def add_book(self, request, pk=None):
        # validate request
        is_valid, error_message = validate_add_book_request(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))
        book_level_name = request.data.get('level')
        book_level = BookLevel.objects.filter(name=book_level_name).first()
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist:
            return Response(status=404)

        if not (request_user_belongs_to_ngo(request, ngo) or get_group_type_from_request_user(
                request.user) != GroupType.NGO_ADMIN):
            return Response(status=403, data=ERROR_403_JSON())
        data = request.data.copy()

        if book_level:
            data['book_level_id'] = book_level.id
        else:
            data['book_level_id'] = None

        data['ngo_id'] = ngo.id

        serializer = BookSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(status=400, data=create_serializer_error(serializer))

    @action(detail=True, methods=['POST'], permission_classes=[CanAddSchool], schema=NGOSchema.add_school())
    def add_school(self, request, pk=None):
        # validate request
        is_valid, error_message = validate_add_school_request(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        school_category_key = request.data.get('school_category')
        school_type_key = request.data.get('school_type')
        medium_key = request.data.get('medium')
        try:
            ngo = NGO.objects.get(key=pk)
            if not (request_user_belongs_to_ngo(request, ngo) or get_group_type_from_request_user(
                    request.user) != GroupType.NGO_ADMIN):
                return Response(status=403, data=ERROR_403_JSON())

            school_category = SchoolCategory.objects.get(name=school_category_key)
            school_type = SchoolType.objects.get(name=school_type_key)
            medium = SchoolMedium.objects.get(name=medium_key)
        except (NGO.DoesNotExist, SchoolCategory.DoesNotExist, SchoolType.DoesNotExist, SchoolMedium.DoesNotExist) as e:
            return Response(status=404, data=create_response_error(e))

        data = request.data.copy()
        data['ngo_key'] = pk
        data['school_category_id'] = school_category.id
        data['school_type_id'] = school_type.id
        data['medium_id'] = medium.id

        if data.get('ward_number') is None or data.get('ward_number') == "":
            data['ward_number'] = None

        if data.get('organization_name') is None or data.get('organization_name') == "":
            data['organization_name'] = None

        if data.get('school_number') is None or data.get('school_number') == "":
            data['school_number'] = None

        if data.get('year_of_intervention') is None or data.get('year_of_intervention') == "":
            data['year_of_intervention'] = None

        serializer = SchoolSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(status=400, data=create_serializer_error(serializer))

    @action(detail=True, methods=['POST'], permission_classes=[CanAddUser], schema=NGOSchema.add_ngo_user())
    def add_user(self, request, pk=None):

        user_type = request.data.get('user_type')
        is_valid, error = validate_user_type(user_type)
        if not is_valid:
            return Response(status=400, data=create_response_data(error))

        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist:
            return Response(status=404)

        if not (request_user_belongs_to_ngo(request, ngo) or get_group_type_from_request_user(
                request.user) == GroupType.READ_ADMIN):
            return Response(status=403, data=ERROR_403_JSON())

        data = request.data.copy()
        data['ngo_id'] = ngo.id
        if data.get('email') is None or data.get('email') == "":
            data['email'] = None
        serializer = UserSerializer(data=data)
        group_name = GroupType(user_type)
        if serializer.is_valid():
            with transaction.atomic():
                ngo_admin_group_name = get_ngo_specific_group_name(group_name, ngo.key)
                # Adding user to group using user's type
                group = Group.objects.get(name=ngo_admin_group_name)
                user = serializer.save()
                user.set_password(data.get('password'))
                user.save()
                user.groups.add(group)
                return Response(serializer.data)
        else:
            return Response(status=400, data=create_serializer_error(serializer))

    @action(detail=True, methods=['GET'])
    def schools(self, request, pk=None):
        sort = request.GET.get('sort')
        schools = School.objects.filter(ngo__key=pk, is_active=True).order_by(sort)
        paginator = pagination.PageNumberPagination()
        result = paginator.paginate_queryset(schools, request)
        serializer = SchoolSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['GET'])
    def get_schools_dropdown(self, request, pk=None):
        schools = School.objects.filter(ngo__key=pk, is_active=True)
        serializer = SchoolSerializer(schools, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['GET'])
    def books(self, request, pk=None):
        sort = request.GET.get('sort')
        books = Book.objects.filter(ngo__key=pk, is_active=True).order_by(sort)
        paginator = pagination.PageNumberPagination()
        result = paginator.paginate_queryset(books, request)
        serializer = BookSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['GET'])
    def mobile_books(self, request, pk=None):
        books = Book.objects.filter(ngo__key=pk, is_active=True)
        serializer = BookSerializer(books, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['GET'])
    def users(self, request, pk=None):
        sort_by = request.GET.get('sort')
        sort = user_sort_by_value(sort_by)
        users = User.objects.filter(ngo__key=pk, is_active=True).order_by(sort)
        paginator = pagination.PageNumberPagination()
        result = paginator.paginate_queryset(users, request)
        serializer = UserSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['POST'], permission_classes=[CanAddClassroom],
            schema=NGOSchema.add_classroom())
    def add_classroom(self, request, pk=None):
        is_valid, error = validate_add_classroom_request(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error))
        standard_name = request.data.get('standard')
        school_key = request.data.get('school')
        try:
            ngo = NGO.objects.get(key=pk)
            if not request_user_belongs_to_ngo(request, ngo):
                return Response(status=403, data=ERROR_403_JSON())

            standard = Standard.objects.get(name=standard_name)
            school = School.objects.get(key=school_key, ngo__key=ngo.key)
        except (NGO.DoesNotExist, Standard.DoesNotExist, School.DoesNotExist):
            return Response(status=404)

        data = request.data.copy()
        data['school_key'] = school.key
        data['standard_id'] = standard.id
        serializer = ClassroomSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        else:
            return Response(status=400, data=create_serializer_error(serializer))

    @action(methods=['GET'], detail=True)
    def get_school_classrooms(self, request, pk=None):
        school_key = request.GET.get('school')
        try:
            school = School.objects.get(key=school_key, ngo__key=pk)
        except School.DoesNotExist:
            return Response(status=404)

        classrooms = Classroom.objects.filter(school=school, is_active=True)
        serializer = ClassroomTableSerializer(classrooms, many=True)
        return Response(serializer.data)

    @action(methods=['GET'], detail=True)
    def students(self, request, pk=None):
        sort = request.GET.get('sort')
        order = request.GET.get('order')
        sort_by = student_sort_by_value(sort, order)
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        common_filters = {
            'classroom__is_active': True,
            'classroom__school__is_active': True,
            'classroom__school__ngo__key': pk,
            'classroom__school__ngo__is_active': True
        }
        paginator = pagination.PageNumberPagination()
        classroom_academic_year = ClassroomAcademicYear.objects.filter(**common_filters, student__is_active=True) \
            .order_by(sort_by)
        result = paginator.paginate_queryset(classroom_academic_year, request)
        serializer = ClassroomAcademicYearSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(methods=['GET'], detail=True)
    def book_fairies(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist:
            return Response(status=404)
        book_fairy_group_name = get_ngo_specific_group_name(GroupType.BOOK_FAIRY, ngo.key)
        fairies = User.objects.filter(ngo__key=pk, groups__name=book_fairy_group_name, is_active=True)
        serializer = UserSerializer(fairies, many=True)
        return Response(serializer.data)

    @action(methods=['POST'], detail=True, permission_classes=[CanAddReadSession])
    def add_session(self, request, pk=None):
        is_valid, error = validate_add_session_request(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error))

        academic_year_key = request.data.get('academic_year')
        try:
            ngo = NGO.objects.get(key=pk)
            academic_year = AcademicYear.objects.get(key=academic_year_key)
        except (NGO.DoesNotExist, AcademicYear.DoesNotExist) as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        classroom_keys = json.loads(request.data.get('classrooms'))
        book_fairy_keys = json.loads(request.data.get('fairies'))
        read_session_dates = json.loads(request.data.get('dates'))
        read_session_data = request.data.copy()
        try:
            with transaction.atomic():
                for date in read_session_dates:
                    read_session_data["academic_year_id"] = academic_year.id
                    start_date_time = date.get('start_date_time')
                    end_date_time = date.get('end_date_time')
                    read_session_data["start_date_time"] = start_date_time
                    read_session_data["end_date_time"] = end_date_time

                    session_type = read_session_data.get('type')

                    # Checking if session is already scheduled for given book fairies?
                    is_session_scheduled = ReadSession.objects \
                        .filter((Q(start_date_time__gt=start_date_time) & Q(start_date_time__lt=end_date_time))
                                | (Q(end_date_time__gt=start_date_time) & Q(end_date_time__lt=end_date_time))
                                | (Q(start_date_time__lt=start_date_time) & Q(end_date_time__gt=start_date_time))
                                | (Q(start_date_time__lt=end_date_time) & Q(end_date_time__gt=end_date_time))
                                | (Q(start_date_time=start_date_time) & Q(end_date_time=end_date_time)),
                                is_cancelled=False,
                                readsessionbookfairy__book_fairy__key__in=book_fairy_keys,
                                academic_year__key=academic_year.key).distinct()

                    if session_type != BOOK_LENDING and is_session_scheduled:
                        error = []
                        for sessions in is_session_scheduled:
                            book_fairy = sessions.readsessionbookfairy_set.all().values_list("book_fairy__first_name",
                                                                                             flat=True)
                            book_fairies = [name for name in book_fairy]
                            error.append({"book_fairy": book_fairies,
                                          "start_date": sessions.start_date_time,
                                          "end_date": sessions.end_date_time})
                        # return Response(status=400, data=create_response_data(error))
                        raise ScheduledBookFairyException(message=create_response_data(error))

                    # Creating Read session
                    read_session_serializer = ReadSessionSerializer(data=read_session_data)
                    if read_session_serializer.is_valid():
                        read_session_serializer.save()
                        read_session_id = read_session_serializer.data.get('id')

                        # Adding read session to multiple classrooms
                        for classroom_key in classroom_keys:
                            classroom = Classroom.objects.get(key=classroom_key)
                            read_session_classroom_data = {"read_session_id": read_session_id,
                                                           "classroom_id": classroom.id}
                            read_session_classroom_serializer = ReadSessionClassroomSerializer(
                                data=read_session_classroom_data)
                            if read_session_classroom_serializer.is_valid():
                                read_session_classroom_serializer.save()
                            else:
                                return Response(status=400,
                                                data=create_serializer_error(read_session_classroom_serializer))

                        # Adding read session to multiple book fairies
                        for book_fairy_key in book_fairy_keys:
                            book_fairy = User.objects.get(key=book_fairy_key)
                            read_session_book_fairy_data = {"read_session_id": read_session_id,
                                                            "user_id": book_fairy.id}
                            read_session_book_fairy_serializer = ReadSessionBookFairySerializer(
                                data=read_session_book_fairy_data)
                            if read_session_book_fairy_serializer.is_valid():
                                read_session_book_fairy_serializer.save()
                            else:
                                return Response(status=400,
                                                data=create_serializer_error(read_session_book_fairy_serializer))
                    else:
                        return Response(status=400, data=create_serializer_error(read_session_serializer))
                return Response(status=201)
        except ScheduledBookFairyException as e:
            return Response(status=400, data=e.message)
        except (Classroom.DoesNotExist, User.DoesNotExist, Exception) as e:
            return Response(status=400, data=create_response_error(e))

    @action(methods=['POST'], detail=True)
    def search(self, request, pk=None):
        try:
            NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        filters = json.loads(request.data.get('search', None))
        model = request.data.get('model', None)
        result = switch(model)(filters, pk, request)
        return result

    @action(detail=True, methods=['GET'])
    def sessions(self, request, pk=None):
        sort = request.GET.get('sort')
        order = request.GET.get('order')
        order_by = session_sort_by_value(sort, order)
        try:
            NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        read_sessions = ReadSession.objects.filter(readsessionclassroom__classroom__school__ngo__key=pk).order_by(
            order_by).distinct()
        paginator = pagination.PageNumberPagination()
        result = paginator.paginate_queryset(read_sessions, request)
        serializer = ReadSessionSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['GET'])
    def get_supervisor_sessions(self, request, pk=None):
        session_type = request.GET.get('type', None)
        sort = request.GET.get('sort')
        order = request.GET.get('order')
        order_by = session_sort_by_value(sort, order)

        try:
            NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        # Get book fairies under supervisor
        book_fairies_keys = SupervisorBookFairy.objects.filter(supervisor=request.user).values_list("book_fairy__key",
                                                                                                    flat=True)
        common_filters = {
            "readsessionbookfairy__book_fairy__key__in": book_fairies_keys,
            "is_cancelled": False
        }

        # Supervisor sessions
        if session_type == SESSION_EVALUATED:
            read_sessions = ReadSession.objects \
                .filter((Q(studentevaluations__read_session__readsessionclassroom__classroom__school__ngo__key=pk) |
                         Q(studentfeedback__read_session__readsessionclassroom__classroom__school__ngo__key=pk) |
                         Q(
                             readsessionhomelendingbook__read_session__readsessionclassroom__classroom__school__ngo__key=pk)),
                        is_evaluated=True, is_verified=False,
                        **common_filters
                        ).order_by(order_by).distinct()

        # Supervisor sessions
        if session_type == SESSION_NON_EVALUATED:
            read_sessions = ReadSession.objects \
                .filter(is_evaluated=False, start_date_time__lte=datetime.now(tz=timezone.utc),
                        readsessionclassroom__classroom__school__ngo__key=pk,
                        **common_filters) \
                .exclude(studentevaluations__read_session__readsessionclassroom__classroom__school__ngo__key=pk,
                         studentfeedback__read_session__readsessionclassroom__classroom__school__ngo__key=pk,
                         readsessionhomelendingbook__read_session__readsessionclassroom__classroom__school__ngo__key=pk) \
                .order_by(order_by) \
                .distinct()

        # All Sessions
        if session_type is None:
            read_sessions = ReadSession.objects.filter(readsessionclassroom__classroom__school__ngo__key=pk,
                                                       **common_filters) \
                .order_by(order_by).distinct()

        paginator = pagination.PageNumberPagination()
        result = paginator.paginate_queryset(read_sessions, request)
        serializer = ReadSessionSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['GET'])
    def book_fairy_sessions(self, request, pk=None):
        book_fairy_key = request.user.key
        session_type = request.GET.get('type', None)
        sort = request.GET.get('sort')
        order = request.GET.get('order')
        order_by = session_sort_by_value(sort, order)

        try:
            ngo = NGO.objects.get(key=pk)
            User.objects.get(key=book_fairy_key, ngo=ngo)
        except (NGO.DoesNotExist, User.DoesNotExist)as e:
            return Response(status=404, data=create_response_error(e))

        common_filter = {
            "readsessionclassroom__classroom__school__ngo__key": pk,
            "readsessionbookfairy__book_fairy__key": book_fairy_key,
        }

        now = datetime.now(tz=timezone.utc)
        if session_type == READ_SESSION_PENDING:
            read_sessions = ReadSession.objects.filter(**common_filter,
                                                       is_verified=False,
                                                       is_evaluated=False,
                                                       start_date_time__lte=now) \
                .order_by(order_by).distinct()
        if session_type == READ_SESSION_EVALUATED_NOT_VERIFIED:
            read_sessions = ReadSession.objects.filter(**common_filter,
                                                       is_verified=False,
                                                       is_evaluated=True,
                                                       start_date_time__lte=now) \
                .order_by(order_by).distinct()

        if session_type == READ_SESSION_UPCOMING:
            read_sessions = ReadSession.objects.filter(**common_filter,
                                                       is_verified=False,
                                                       is_evaluated=False,
                                                       start_date_time__gte=now) \
                .order_by(order_by).distinct()

        paginator = pagination.PageNumberPagination()
        result = paginator.paginate_queryset(read_sessions, request)
        serializer = ReadSessionSerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['GET'])
    def mobile_book_fairy_sessions(self, request, pk=None):
        book_fairy_key = request.user.key
        session_type = request.GET.get('type', None)

        try:
            ngo = NGO.objects.get(key=pk)
            User.objects.get(key=book_fairy_key, ngo=ngo)
        except (NGO.DoesNotExist, User.DoesNotExist)as e:
            return Response(status=404, data=create_response_error(e))

        common_filter = {
            "readsessionclassroom__classroom__school__ngo__key": pk,
            "readsessionbookfairy__book_fairy__key": book_fairy_key,
        }

        now = datetime.now(tz=timezone.utc)
        if session_type == READ_SESSION_PENDING:
            read_sessions = ReadSession.objects.filter(**common_filter,
                                                       is_verified=False,
                                                       is_evaluated=False,
                                                       start_date_time__lte=now) \
                .order_by("start_date_time").distinct()
        elif session_type == READ_SESSION_EVALUATED_NOT_VERIFIED:
            read_sessions = ReadSession.objects.filter(**common_filter,
                                                       is_verified=False,
                                                       is_evaluated=True,
                                                       start_date_time__lte=now) \
                .order_by("start_date_time").distinct()

        elif session_type == READ_SESSION_UPCOMING:
            a_week_later = now + timedelta(days=7)
            read_sessions = ReadSession.objects.filter(**common_filter,
                                                       is_verified=False,
                                                       is_evaluated=False,
                                                       start_date_time__gte=now,
                                                       end_date_time__lte=a_week_later) \
                .order_by("start_date_time").distinct()

        else:
            read_sessions = ReadSession.objects.filter(**common_filter) \
                .order_by("start_date_time").distinct()

        serializer = ReadSessionSerializer(read_sessions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['POST'], permission_classes=[CanDeleteNGO])
    def deactivate_ngos(self, request, pk=None):
        # TODO activate ngos??
        keys = request.data.get('keys')
        ngo_keys = json.loads(keys)
        ngos = NGO.objects.filter(key__in=ngo_keys)
        for ngo in ngos:
            ngo.is_active = False
            ngo.save()
        return Response(status=200)

    @action(methods=['POST'], detail=True, schema=UserSchema.file_import(), permission_classes=[CanImportUsers])
    def import_users(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        is_valid, error_message = validate_user_file_import(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        file_in_memory = request.FILES['file'].read()
        workbook = load_workbook(filename=BytesIO(file_in_memory))
        try:
            worksheet = workbook[USER_WORKSHEET_NAME]
        except KeyError:
            return Response(status=400, data=create_response_data("Invalid excel file"))

        rows = list(worksheet.rows)
        is_valid, error_message = validate_user_file_excel_content(rows)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        response = []
        try:
            with transaction.atomic():
                error_in_file = False
                for index, row in enumerate(rows):
                    if index == 0:
                        continue
                    key = str(row[0].value).strip() if row[0].value else None
                    username = str(row[1].value).strip() if row[1].value else None
                    first_name = str(row[2].value).strip() if row[2].value else None
                    middle_name = str(row[3].value).strip() if row[3].value else None
                    last_name = str(row[4].value).strip() if row[4].value else None
                    email = str(row[5].value).strip() if row[5].value else None
                    role = str(row[6].value).strip() if row[6].value else None

                    if username is None and first_name is None and last_name is None and email is None and role is None:
                        continue

                    if key:
                        # Check if user with key exists
                        existing_user = User.objects.filter(key=key, username=username, ngo=ngo).first()
                        if existing_user:
                            group_type = get_group_type_from_name(role)
                            if group_type:
                                user_data = {"first_name": first_name, "middle_name": middle_name,
                                             "last_name": last_name, "ngo_id": ngo.id, "email": email,
                                             "username": username, "is_active": existing_user.is_active}

                                if email is None and group_type != GroupType.BOOK_FAIRY:
                                    response.append(create_file_upload_error(index, "email", "Email is required"))
                                    error_in_file = True

                                serializer = UserSerializer(existing_user, data=user_data)
                                if serializer.is_valid():
                                    new_group_name = get_ngo_specific_group_name(group_type, ngo.key)
                                    new_group = Group.objects.get(name=new_group_name)
                                    user = serializer.save()
                                    user.save()
                                    user.groups.clear()
                                    user.groups.add(new_group)
                                    logger.debug("Saving existing user :" + str(index))
                                else:
                                    response.append(
                                        create_file_upload_serializer_error(index, serializer.errors))
                                    error_in_file = True
                            else:
                                logger.error("Incorrect user group provided role")
                                response.append(
                                    create_file_upload_error(index, 'group', 'Value must be one of ' + ', '.join(
                                        get_valid_user_types())))
                                error_in_file = True

                        else:
                            logger.error("User does not exist index: " + str(index))
                            response.append(create_file_upload_error(index, 'key', 'User with key does not exist'))
                            error_in_file = True
                    else:
                        new_group_type = get_group_type_from_name(role)
                        if new_group_type:
                            new_user = {"username": username, "first_name": first_name, "middle_name": middle_name,
                                        "last_name": last_name, "email": email, "ngo_id": ngo.id}
                            if email is None and new_group_type != GroupType.BOOK_FAIRY:
                                response.append(create_file_upload_error(index, "email", "Email is required"))
                                error_in_file = True

                            serializer = UserSerializer(data=new_user)
                            if serializer.is_valid():
                                user = serializer.save()
                                user.set_password("admin")
                                user.save()
                                new_group_name = get_ngo_specific_group_name(new_group_type, ngo.key)
                                new_group = Group.objects.get(name=new_group_name)
                                user.groups.add(new_group)
                            else:
                                response.append(create_file_upload_serializer_error(index, serializer.errors))
                                error_in_file = True

                        else:
                            # Incorrect role /group name
                            error_in_file = True
                            logger.error("Incorrect user group provided role: %s", role)
                            response.append(
                                create_file_upload_error(index, 'group', 'Value must be one of ' + ', '.join(
                                    get_valid_user_types())))
                if error_in_file:
                    raise DatabaseError
        except DatabaseError as e:
            logger.error("Error in import users excel file")
            if len(response) == 0:
                return Response(status=500, data=create_response_error(e))
            pass
        return Response(status=200, data=create_response_data(response))

    @action(methods=['GET'], detail=True, permission_classes=[CanExportUsers])
    def export_users(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        users = User.objects.filter(ngo=ngo, is_active=True)
        excel_data = write_to_excel(FileType.USER, users)
        base64_excel_data = base64.b64encode(excel_data)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=Users.xlsx'
        response.write(base64_excel_data)
        return response

    @action(methods=['GET'], detail=True)
    def get_levels(self, request, pk=None):
        try:
            NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        levels = Level.objects.filter(ngo__key=pk)
        serializer = LevelSerializer(levels, many=True)
        return Response(serializer.data)

    @action(methods=['GET'], detail=True)
    def get_read_session_books(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        query_set = Book.objects.filter(ngo=ngo)
        serializer = BookSerializer(query_set, many=True)
        return Response(serializer.data)

    @action(methods=['POST'], detail=True, permission_classes=[CanImportBook])
    def import_books(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        is_valid, error_message = validate_user_file_import(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        file_in_memory = request.FILES['file'].read()
        workbook = load_workbook(filename=BytesIO(file_in_memory))
        worksheet = workbook[BOOK_WORKSHEET_NAME]
        rows = list(worksheet.rows)
        is_valid, error_message = validate_book_file_excel_content(rows)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        book_levels = BookLevel.objects.all()
        response = []
        try:
            with transaction.atomic():
                error_in_file = False
                for index, row in enumerate(rows):
                    if index == 0:
                        continue
                    key = row[0].value.strip() if row[0].value else None
                    name = row[1].value.strip() if row[1].value else None
                    book_level_name = row[2].value.strip() if row[2].value else None

                    if name is None and key is None and book_level_name is None:
                        continue

                    book_level = get_book_level(book_levels, book_level_name)
                    if not book_level and book_level_name:
                        response.append(create_file_upload_error(index, 'level', 'Value must be one of ' + str(
                            get_valid_book_levels())))
                        error_in_file = True
                        continue
                    author = row[3].value.strip() if row[3].value else None
                    publisher = row[4].value.strip() if row[4].value else None
                    price = str(row[5].value) if row[5].value else None
                    if key:
                        # Check if book with key exists
                        existing_book = Book.objects.filter(key=key, ngo=ngo).first()
                        if existing_book:
                            book_data = {"name": name, "publisher": publisher, "price": price, "ngo_id": ngo.id,
                                         "author": author,
                                         "is_active": existing_book.is_active}
                            if book_level:
                                book_data['book_level_id'] = book_level.id
                            serializer = BookSerializer(existing_book, data=book_data)
                            if serializer.is_valid():
                                serializer.save()
                            else:
                                response.append(create_file_upload_serializer_error(index, serializer.errors))
                                error_in_file = True
                        else:
                            logging.error("Book with specified key does not exist index:" + str(index))
                            response.append(create_file_upload_error(index, "book",
                                                                     "Book with specified key does not exist"))
                            error_in_file = True
                    else:
                        existing_book = Book.objects.filter(name=name, ngo=ngo).first()
                        if existing_book:
                            response.append(create_file_upload_error(index, "name", "Book with same name already "
                                                                                    "exists"))
                            error_in_file = True
                            continue

                        new_book_data = {"name": name, "publisher": publisher, "price": price, "ngo_id": ngo.id,
                                         "author": author}
                        if book_level:
                            new_book_data['book_level_id'] = book_level.id
                        serializer = BookSerializer(data=new_book_data)
                        if serializer.is_valid():
                            serializer.save()
                            logging.debug("Creating new book index: " + str(index))
                        else:
                            response.append(create_file_upload_serializer_error(index, serializer.errors))
                            error_in_file = True

                if error_in_file:
                    raise DatabaseError
        except DatabaseError:
            logging.error("Error in import books excel file")
            pass
        return Response(data=create_response_data(response))

    @action(methods=['POST'], detail=True, permission_classes=[CanImportBook])
    def import_books_to_inventory(self, request, pk=None):
        book_key = request.data.get('book')
        try:
            ngo = NGO.objects.get(key=pk)
            book = Book.objects.get(key=book_key)
        except (NGO.DoesNotExist, Book.DoesNotExist) as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        is_valid, error_message = validate_user_file_import(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        file_in_memory = request.FILES['file'].read()
        workbook = load_workbook(filename=BytesIO(file_in_memory))
        worksheet = workbook[INVENTORY_BOOK_WORKSHEET_NAME]
        rows = list(worksheet.rows)
        is_valid, error_message = validate_inventory_book_file_excel_content(rows)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        response = []
        try:
            with transaction.atomic():
                error_in_file = False
                for index, row in enumerate(rows):
                    if index == 0:
                        continue
                    key = str(row[0].value).strip() if row[0].value else None
                    serial_number = str(row[1].value).strip() if row[1].value else None
                    status_string = str(row[2].value).strip() if row[2].value else None
                    status = get_inventory_status(status_string)
                    year_of_purchase_string = row[3].value
                    year_of_purchase = None
                    if year_of_purchase_string:
                        try:
                            if type(year_of_purchase_string) is not datetime:
                                year_of_purchase_string = str(row[3].value).strip() if row[3].value else None
                                year_of_purchase = datetime.strptime(year_of_purchase_string, "%Y").date()
                            else:
                                year_of_purchase = year_of_purchase_string.date()
                        except ValueError:
                            error_in_file = True
                            response.append(
                                create_file_upload_error(index, 'year_of_intervention', 'Value must be yyyy'))
                            continue

                    if serial_number is None and status is None:
                        continue

                    if not status:
                        logger.error("Inventory status is incorrect: index" + str(index))
                        response.append(create_file_upload_error(index, 'status', 'Value must be one of ' + ', '.join(
                            get_valid_inventory_statuses())))
                        error_in_file = True
                        continue

                    if not serial_number:
                        logger.error("Inventory serial number is incorrect: index" + str(index))
                        response.append(create_file_upload_error(index, 'serial number',
                                                                 'Inventory serial number is incorrect'))
                        error_in_file = True
                        continue

                    if key:
                        # Check inventory with key exist?
                        existing_inventory = Inventory.objects.filter(key=key, book=book).first()

                        if existing_inventory:
                            inventory_data = {"status": status, "serial_number": serial_number, "book_id": book.id,
                                              "added_date_time": year_of_purchase,
                                              "is_active": existing_inventory.is_active}

                            serializer = InventorySerializer(existing_inventory, data=inventory_data)
                            if serializer.is_valid():
                                serializer.save()
                            else:
                                response.append(create_file_upload_serializer_error(index, serializer.errors))
                                error_in_file = True

                        else:
                            logging.error("Inventory with specified key does not exist index:" + str(index))
                            response.append(create_file_upload_error(index, "inventory",
                                                                     "Inventory with specified key does not exist"))
                            error_in_file = True
                    else:
                        existing_inventory_with_same_serial_number = Inventory.objects.filter(book__ngo=ngo,
                                                                                              serial_number=serial_number).exists()
                        if existing_inventory_with_same_serial_number:
                            response.append(create_file_upload_error(index, 'serial_number',
                                                                     "An inventory exists with the same serial_number :" + serial_number))
                            error_in_file = True
                            continue

                        new_inventory_data = {"status": status, "serial_number": serial_number, "book_id": book.id,
                                              "added_date_time": year_of_purchase}
                        serializer = InventorySerializer(data=new_inventory_data)
                        if serializer.is_valid():
                            serializer.save()
                            logging.debug("Creating new inventory index: " + str(index))
                        else:
                            response.append(create_file_upload_serializer_error(index, serializer.errors))
                            error_in_file = True

                if error_in_file:
                    raise DatabaseError
        except DatabaseError:
            logging.error("Error in import books in inventory excel file")
            pass
        return Response(status=200, data=create_response_data(response))

    @action(methods=['POST'], detail=True, permission_classes=[CanImportBook])
    def import_books_and_create_inventory(self, request, pk=None):
        return Response(status=404)
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        is_valid, error_message = validate_user_file_import(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        file_in_memory = request.FILES['file'].read()
        workbook = load_workbook(filename=BytesIO(file_in_memory))
        worksheet = workbook[BOOK_WORKSHEET_NAME]
        rows = list(worksheet.rows)
        is_valid, error_message = validate_book_and_inventory_file_excel_content(rows)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        book_levels = BookLevel.objects.all()
        response = []
        try:
            with transaction.atomic():
                error_in_file = False
                for index, row in enumerate(rows):
                    if index == 0:
                        continue
                    key = row[0].value.strip() if row[0].value else None
                    name = row[1].value.strip() if row[1].value else None
                    book_level_name = row[2].value.strip() if row[2].value else None

                    if name is None and key is None and book_level_name is None:
                        continue

                    book_level = get_book_level(book_levels, book_level_name)
                    if not book_level and book_level_name:
                        response.append(create_file_upload_error(index, 'level', 'Value must be one of ' + str(
                            get_valid_book_levels())))
                        error_in_file = True
                        continue
                    author = row[3].value.strip() if row[3].value else None
                    publisher = row[4].value.strip() if row[4].value else None
                    price = str(row[5].value) if row[5].value else None
                    serial = str(row[6].value).strip() if row[6].value else None
                    copies = int(str(row[7].value).strip()) if row[7].value else None

                    existing_book = Book.objects.filter(name=name, ngo=ngo).first()
                    if existing_book:

                        book_data = {"name": name, "publisher": publisher, "price": price, "ngo_id": ngo.id,
                                     "author": author,
                                     "is_active": True}
                        if book_level:
                            book_data['book_level_id'] = book_level.id
                        serializer = BookSerializer(existing_book, data=book_data)
                        if serializer.is_valid():
                            serializer.save()
                        else:
                            response.append(create_file_upload_serializer_error(index, serializer.errors))
                            error_in_file = True

                        # Create inventory
                        if not serial:
                            response.append(create_file_upload_error(index, 'serial', 'blank serial'))
                            error_in_file = True

                        if not copies:
                            response.append(create_file_upload_error(index, 'copies', 'blank copies'))
                            error_in_file = True

                        if not copies or copies < 1:
                            response.append(create_file_upload_error(index, 'copies', 'invalid copies'))
                            error_in_file = True

                        # print(serial)
                        # print(serial.find(".1", -2))

                        #  FOR PMC
                        if serial.find(".1", -2) == -1:
                            response.append(create_file_upload_error(index, 'serial', '.1 is missing serial'))
                            error_in_file = True
                            continue

                        #  FOR PCMC
                        # if serial.find("-0001", -5) == -1:
                        #     response.append(create_file_upload_error(index, 'serial', '-0001 is missing serial'))
                        #     error_in_file = True
                        #     continue

                        # FOR PMC
                        serial_base = serial[:-1]
                        for x in range(1, copies + 1):
                            serial_number = serial_base + str(x)
                            error_in_file, errors = create_inventory(existing_book, serial_number, ngo)
                            if error_in_file:
                                response.append(create_file_upload_serializer_error(index, errors))

                        # FOR PCMC
                        # serial_base = serial[:-5]
                        # for x in range(1, copies + 1):
                        #     serial_number = serial_base + '-' + ("%04d" % x)
                        #     # print(serial_number)
                        #     error_in_file, errors = create_inventory(existing_book, serial_number, ngo)
                        #     if error_in_file:
                        #         response.append(create_file_upload_serializer_error(index, errors))


                    else:
                        new_book_data = {"name": name, "publisher": publisher,
                                         "price": price, "ngo_id": ngo.id,
                                         "author": author}
                        if book_level:
                            new_book_data['book_level_id'] = book_level.id
                        serializer = BookSerializer(data=new_book_data)
                        if serializer.is_valid():
                            serializer.save()
                            # logging.debug("Creating new book index: " + str(index))
                        else:
                            response.append(create_file_upload_serializer_error(index, serializer.errors))
                            error_in_file = True

                if error_in_file:
                    raise DatabaseError
        except DatabaseError:
            logging.error("Error in import books and inventory excel file")
            pass
        return Response(data=create_response_data(response))

    @action(methods=['GET'], detail=True, permission_classes=[CanExportBook])
    def export_books(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        books = Book.objects.filter(ngo=ngo, is_active=True)
        excel_data = write_to_excel(FileType.BOOK, books)
        base64_excel_data = base64.b64encode(excel_data)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + BOOK_WORKSHEET_NAME + '.xlsx'
        response.write(base64_excel_data)
        return response

    @action(methods=['GET'], detail=True, permission_classes=[CanExportBook])
    def export_inventory(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist  as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        book_key = request.GET.get('book')
        try:
            Book.objects.get(key=book_key, ngo=ngo)
        except Book.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        inventory = Inventory.objects.filter(book__key=book_key, is_active=True)
        excel_data = write_to_excel(FileType.INVENTORY, inventory)
        base64_excel_data = base64.b64encode(excel_data)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + INVENTORY_BOOK_WORKSHEET_NAME + '.xlsx'
        response.write(base64_excel_data)
        return response

    @action(methods=['POST'], detail=True, permission_classes=[CanImportSchool])
    def import_schools(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        is_valid, error_message = validate_user_file_import(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        file_in_memory = request.FILES['file'].read()
        workbook = load_workbook(filename=BytesIO(file_in_memory))
        worksheet = workbook[SCHOOL_WORKSHEET_NAME]
        rows = list(worksheet.rows)
        is_valid, error_message = validate_school_file_excel_content(rows)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        school_categories = SchoolCategory.objects.all()
        school_types = SchoolType.objects.all()
        mediums = SchoolMedium.objects.all()
        response = []
        try:
            with transaction.atomic():
                error_in_file = False
                for index, row in enumerate(rows):
                    if index == 0:
                        continue
                    key = str(row[0].value).strip() if row[0].value else None
                    name = str(row[1].value).strip() if row[1].value else None
                    address = str(row[2].value).strip() if row[2].value else None
                    pin_code = row[3].value
                    ward_number = str(row[4].value).strip() if row[4].value else None
                    school_number = str(row[5].value).strip() if row[5].value else None
                    school_category_name = str(row[6].value).strip() if row[6].value else None
                    school_type_name = str(row[7].value).strip() if row[7].value else None
                    medium_name = str(row[8].value).strip() if row[8].value else None
                    organization_name = str(row[9].value).strip() if row[9].value else None
                    year_of_intervention_string = row[10].value
                    year_of_intervention = None

                    if name is None and address is None and pin_code is None and school_category_name is None and \
                            school_type_name is None and medium_name is None:
                        continue

                    school_category = get_school_category(school_categories, school_category_name)
                    if not school_category:
                        error_in_file = True
                        response.append(create_file_upload_error(index, 'school_category',
                                                                 'Value must be one of ' + ', '.join(
                                                                     get_valid_school_categories())))
                        continue

                    school_type = get_school_type(school_types, school_type_name)
                    if not school_type:
                        error_in_file = True
                        response.append(create_file_upload_error(index, 'school_type',
                                                                 'Value must be one of ' + ', '.join(
                                                                     get_valid_school_types())))
                        continue

                    medium = get_school_medium(mediums, medium_name)
                    if not medium:
                        error_in_file = True
                        response.append(create_file_upload_error(index, 'medium', 'Value must be one of ' + ', '.join(
                            get_valid_school_mediums())))
                        continue
                    if year_of_intervention_string:
                        try:
                            if type(year_of_intervention_string) is not datetime:
                                year_of_intervention_string = str(row[10].value).strip() if row[10].value else None
                                year_of_intervention = datetime.strptime(year_of_intervention_string, "%Y").date()
                            else:
                                year_of_intervention = year_of_intervention_string.date()
                        except ValueError:
                            error_in_file = True
                            response.append(
                                create_file_upload_error(index, 'year_of_intervention', 'Value must be yyyy'))
                            continue

                    if key:
                        # Check if school with key exists
                        existing_school = School.objects.filter(key=key, ngo=ngo).first()
                        if existing_school:
                            school_data = {"name": name, "address": address, "pin_code": pin_code,
                                           "ward_number": ward_number, "school_number": school_number,
                                           "school_category": school_category, "school_type": school_type,
                                           "medium": medium, "organization_name": organization_name,
                                           "year_of_intervention": year_of_intervention, 'ngo_key': pk,
                                           'school_category_id': school_category.id, 'school_type_id': school_type.id,
                                           'medium_id': medium.id,
                                           'is_active': existing_school.is_active}
                            serializer = SchoolSerializer(existing_school, data=school_data)
                            if serializer.is_valid():
                                serializer.save()
                            else:
                                response.append(create_file_upload_serializer_error(index, serializer.errors))
                                error_in_file = True
                        else:
                            logging.error("School with specified key does not exist index:" + str(index))
                            response.append(create_file_upload_error(index, "school",
                                                                     "School with specified key does not exist"))
                            error_in_file = True
                    else:

                        existing_school = School.objects.filter(name=name, ngo=ngo).first()
                        if existing_school:
                            response.append(create_file_upload_error(index, "name", "School with same name already "
                                                                                    "exists"))
                            error_in_file = True
                            continue
                        new_school = {"name": name, "address": address, "pin_code": pin_code,
                                      "ward_number": ward_number, "school_number": school_number,
                                      "school_category": school_category, "school_type": school_type, "medium": medium,
                                      "organization_name": organization_name,
                                      "year_of_intervention": year_of_intervention, 'ngo_key': pk,
                                      'school_category_id': school_category.id, 'school_type_id': school_type.id,
                                      'medium_id': medium.id}
                        serializer = SchoolSerializer(data=new_school)
                        if serializer.is_valid():
                            serializer.save()
                            logging.debug("Creating new school index: " + str(index))
                        else:
                            response.append(create_file_upload_serializer_error(index, serializer.errors))
                            error_in_file = True

                if error_in_file:
                    raise DatabaseError
        except DatabaseError:
            logging.error("Error in import schools excel file")
            pass
        return Response(data=create_response_data(response))

    @action(methods=['GET'], detail=True, permission_classes=[CanExportSchool])
    def export_schools(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        schools = School.objects.filter(ngo=ngo, is_active=True)
        excel_data = write_to_excel(FileType.SCHOOL, schools)
        base64_excel_data = base64.b64encode(excel_data)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + SCHOOL_WORKSHEET_NAME + '.xlsx'
        response.write(base64_excel_data)
        return response

    @action(detail=True, methods=['POST'], permission_classes=[CanAddLevel])
    def add_level(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_ngo(request, ngo):
            return Response(status=403, data=ERROR_403_JSON())

        data = request.data.copy()
        data["ngo_id"] = ngo.id
        serializer = LevelSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(status=201)
        else:
            return Response(status=400, data=create_serializer_error(serializer))

    @action(detail=True, methods=['GET'])
    def supervisors(self, request, pk=None):
        try:
            ngo = NGO.objects.get(key=pk)
        except NGO.DoesNotExist  as e:
            return Response(status=404, data=create_response_error(e))
        supervisor_group_name = get_ngo_specific_group_name(GroupType.SUPERVISOR, ngo.key)
        supervisors = User.objects.filter(ngo__key=pk, groups__name=supervisor_group_name, is_active=True)
        serializer = UserSerializer(supervisors, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['GET'])
    def get_book_fairy_under_supervisor(self, request, pk=None):
        supervisor_key = request.GET.get('supervisor')
        try:
            ngo = NGO.objects.get(key=pk)
            supervisor = User.objects.get(key=supervisor_key)
        except (NGO.DoesNotExist, User.DoesNotExist) as e:
            return Response(status=404, data=create_response_error(e))

        book_fairies_keys = SupervisorBookFairy.objects.filter(supervisor=supervisor).values_list("book_fairy__key",
                                                                                                  flat=True)
        book_fairies = User.objects.filter(key__in=book_fairies_keys)
        serializer = UserSerializer(book_fairies, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['POST'], permission_classes=[CanChangeNGO])
    def add_book_fairies_under_supervisor(self, request, pk=None):
        supervisor_key = request.data.get('supervisor')

        try:
            ngo = NGO.objects.get(key=pk)
            if not request_user_belongs_to_ngo(request, ngo):
                return Response(status=403, data=ERROR_403_JSON())
            supervisor_group_name = get_ngo_specific_group_name(GroupType.SUPERVISOR, ngo.key)
            supervisor = User.objects.get(key=supervisor_key,
                                          groups__name=supervisor_group_name,
                                          ngo=ngo)
        except (NGO.DoesNotExist, User.DoesNotExist) as e:
            return Response(status=404, data=create_response_error(e))

        book_fairy_keys = json.loads(request.data.get("fairies"))
        try:
            with transaction.atomic():
                # is supervisor present in supervisorbookfairy??
                is_supervisor_present = SupervisorBookFairy.objects.filter(supervisor=supervisor)
                if len(is_supervisor_present) > 0:
                    is_supervisor_present.delete()

                # Add fresh book fairy entries under supervisor
                for key in book_fairy_keys:
                    book_fairy = User.objects.get(ngo=ngo, key=key)
                    supervisor_book_fairy_data = {"supervisor_id": supervisor.id,
                                                  "book_fairy_id": book_fairy.id}
                    serializer = SupervisorBookFairySerializer(data=supervisor_book_fairy_data)
                    if serializer.is_valid():
                        serializer.save()

                return Response(status=204)
        except User.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

    @action(methods=['GET'], detail=True)
    def search_book_by_serial_number(self, request, pk=None):
        serial_no = request.GET.get('serial_number')
        try:
            NGO.objects.get(key=pk)
        except NGO.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        # books = Book.objects.filter(inventory__serial_number__icontains=serial_no,
        #                             ngo__key=pk)
        # serializer = BookSerializer(books, many=True)
        books = Inventory.objects.filter(serial_number__icontains=serial_no,
                                         book__ngo__key=pk,
                                         book__is_active=True,
                                         is_active=True)
        serializer = InventorySerializer(books, many=True)
        return Response(serializer.data)

    @action(methods=['POST'], detail=True)
    def get_student_session_report(self, request, pk=None):
        book_fairy_key = request.data.get('book_fairy')
        academic_year_key = request.data.get('academic_year')
        school_key = request.data.get('school')
        classroom_key = request.data.get('classroom')
        session_type = request.data.get('type')
        attendance = request.data.get('attendance')
        try:
            ngo = NGO.objects.get(key=pk)
            book_fairy = User.objects.get(key=book_fairy_key)
            academic_year = AcademicYear.objects.get(key=academic_year_key)
            school = School.objects.get(ngo=ngo, key=school_key)
            classroom = Classroom.objects.get(school=school, key=classroom_key)
        except (NGO.DoesNotExist, User.DoesNotExist, AcademicYear.DoesNotExist, School.DoesNotExist,
                Classroom.DoesNotExist) as e:
            return Response(status=404)

        filters = {
            'readsessionbookfairy__book_fairy': book_fairy,
            'readsessionclassroom__classroom': classroom,
            'readsessionclassroom__classroom__school': school,
            'academic_year': academic_year,
            'is_evaluated': True
        }
        if session_type == REGULAR or session_type == EVALUATION:
            sessions = ReadSession.objects.exclude(type=BOOK_LENDING) \
                .filter(**filters, type=session_type).order_by("start_date_time")
            serializer = ReadSessionBookFairyReportSerializer(sessions, many=True, attendance=attendance)
        else:
            sessions = ReadSession.objects.exclude(type=BOOK_LENDING) \
                .filter(**filters).order_by("start_date_time")
            serializer = ReadSessionBookFairyReportSerializer(sessions, many=True, attendance=attendance)

        return Response(status=200, data=serializer.data)


class LevelViewSet(ViewSet):

    def list(self, request):
        queryset = Level.objects.all()
        serializer = LevelSerializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, pk=None):
        queryset = Level.objects.all()
        item = get_object_or_404(queryset, pk=pk)
        serializer = LevelSerializer(item)
        return Response(serializer.data)

    def update(self, request, pk=None):
        if not has_permission(request, PERMISSION_CAN_CHANGE_LEVEL):
            return Response(status=403, data=ERROR_403_JSON())

        try:
            level = Level.objects.get(key=pk)
        except Level.DoesNotExist:
            return Response(status=404)

        if not request_user_belongs_to_ngo(request, level.ngo):
            return Response(status=403, data=ERROR_403_JSON())

        data = request.data.copy()
        data['ngo'] = request.user.ngo
        data['rank'] = level.rank
        serializer = LevelSerializer(level, data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(status=400, data=create_serializer_error(serializer))
