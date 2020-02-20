#  Copyright (c) 2020. Maverick Labs
#    This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU Affero General Public License as,
#   published by the Free Software Foundation, either version 3 of the
#   License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#   along with this program.  If not, see <http://www.gnu.org/licenses/>.

import base64
import logging

import pyqrcode
from datetime import datetime
from django.db import DatabaseError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.crypto import get_random_string
from io import BytesIO
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from rest_framework import pagination
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.utils import json
from rest_framework.viewsets import ViewSet

from books.models import Book, Inventory, BookLevel
from books.serializers import BookSerializer, InventorySerializer
from books.validators import validate_inventory_book_file_excel_content, validate_book_and_inventory_file_excel_content
from read.common import convert_to_dropdown
from read.constants import ERROR_403_JSON, BOOK_WORKSHEET_NAME, INVENTORY_BOOK_WORKSHEET_NAME, \
    LENGTH_QR_CODE_SVG_FILENAME
from read.utils import create_response_error, create_serializer_error, request_user_belongs_to_book_ngo, \
    request_user_belongs_to_books_ngo, write_to_pdf, create_response_data, get_inventory_status, \
    create_file_upload_error, get_valid_inventory_statuses, get_book_level, get_valid_book_levels, \
    create_file_upload_serializer_error
from users.permissions import has_permission, PERMISSION_CAN_VIEW_BOOK, PERMISSION_CAN_CHANGE_BOOK, \
    PERMISSION_CAN_CHANGE_INVENTORY, CanDeleteInventory, CanDeleteBook, CanViewInventory
from users.validators import validate_user_file_import

logger = logging.getLogger(__name__)


class BookViewSet(ViewSet):

    def retrieve(self, request, pk=None):
        if not has_permission(request, PERMISSION_CAN_VIEW_BOOK):
            return Response(status=403, data=ERROR_403_JSON())
        queryset = Book.objects.all()
        book = get_object_or_404(queryset, key=pk)

        if not request_user_belongs_to_book_ngo(request, book):
            return Response(status=403, data=ERROR_403_JSON())

        serializer = BookSerializer(book)
        return Response(serializer.data)

    def update(self, request, pk=None):
        if not has_permission(request, PERMISSION_CAN_CHANGE_BOOK):
            return Response(status=403, data=ERROR_403_JSON())
        book_level = None
        book_level_name = request.data.get('level')
        try:
            book = Book.objects.get(key=pk)
            if book_level_name is not None:
                book_level = BookLevel.objects.get(name=book_level_name)
        except (Book.DoesNotExist, BookLevel.DoesNotExist) as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_book_ngo(request, book):
            return Response(status=403, data=ERROR_403_JSON())

        data = request.data.copy()
        if book_level:
            data['book_level_id'] = book_level.id
        else:
            data['book_level_id'] = book.level.id

        data['ngo_id'] = book.ngo.id

        serializer = BookSerializer(book, data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(status=400, data=create_serializer_error(serializer))

    @action(detail=False, methods=['POST'], permission_classes=[CanDeleteBook])
    def deactivate_book(self, request, pk=None):
        keys = request.data.get('keys')
        book_keys = json.loads(keys)

        if not request_user_belongs_to_books_ngo(request, book_keys):
            return Response(status=403, data=ERROR_403_JSON())
        # TODO use atomic transactions
        books = Book.objects.filter(key__in=book_keys, ngo=request.user.ngo)
        for book in books:
            book.is_active = False
            book.save()

        return Response(status=204)

    @action(detail=True, methods=['GET'], permission_classes=[CanViewInventory])
    def inventory(self, request, pk=None):
        order = request.GET.get('sort')
        try:
            book = Book.objects.get(key=pk)
        except Book.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_book_ngo(request, book):
            return Response(status=403, data=ERROR_403_JSON())

        paginator = pagination.PageNumberPagination()
        books = Inventory.objects.filter(book__key=pk, book__is_active=True, is_active=True).order_by(order)
        result = paginator.paginate_queryset(books, request)
        serializer = InventorySerializer(result, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['GET'], permission_classes=[CanViewInventory])
    def mobile_inventory(self, request, pk=None):
        try:
            book = Book.objects.get(key=pk)
        except Book.DoesNotExist as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_book_ngo(request, book):
            return Response(status=403, data=ERROR_403_JSON())

        books = Inventory.objects.filter(book__key=pk, book__is_active=True, is_active=True)
        serializer = InventorySerializer(books, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['POST'], permission_classes=[AllowAny])
    def generate_qr_code(self, request, pk=None):

        is_valid, error_message = validate_user_file_import(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        file_in_memory = request.FILES['file'].read()
        workbook = load_workbook(filename=BytesIO(file_in_memory))
        worksheet = workbook[INVENTORY_BOOK_WORKSHEET_NAME]
        rows = list(worksheet.rows)
        is_valid, error_message = validate_inventory_book_file_excel_content(rows)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        book_name = request.data.get('file_name')
        book_name = str(book_name).split('.')[0] if book_name else None
        if not book_name:
            return Response(status=400, data=create_response_data("Invalid file name"))

        response = []
        serial_numbers = []
        error_in_file = False
        try:
            for index, row in enumerate(rows):
                if index == 0:
                    continue
                key = str(row[0].value).strip() if row[0].value else None
                serial_number = str(row[1].value).strip() if row[1].value else None
                status_string = str(row[2].value).strip() if row[2].value else None
                status = get_inventory_status(status_string)
                year_of_purchase_string = row[3].value
                year_of_purchase = None
                if year_of_purchase_string:
                    try:
                        if type(year_of_purchase_string) is not datetime:
                            year_of_purchase_string = str(row[3].value).strip() if row[3].value else None
                            year_of_purchase = datetime.strptime(year_of_purchase_string, "%Y").date()
                        else:
                            year_of_purchase = year_of_purchase_string.date()
                    except ValueError:
                        error_in_file = True
                        response.append(
                            create_file_upload_error(index, 'year_of_intervention', 'Value must be yyyy'))
                        continue

                if serial_number is None and status is None:
                    continue

                if not status:
                    logger.error("Inventory status is incorrect: index" + str(index))
                    response.append(create_file_upload_error(index, 'status', 'Value must be one of ' + ', '.join(
                        get_valid_inventory_statuses())))
                    error_in_file = True
                    continue

                if not serial_number:
                    logger.error("Inventory serial number is incorrect: index" + str(index))
                    response.append(create_file_upload_error(index, 'serial number',
                                                             'Inventory serial number is incorrect'))
                    error_in_file = True
                    continue

                serial_numbers.append(serial_number)
            if error_in_file:
                raise DatabaseError
        except DatabaseError:
            logging.error("Error in inventory excel file")
            return Response(status=200, data=create_response_data(response))

        buffer = BytesIO()
        my_canvas = canvas.Canvas(buffer, bottomup=0)
        my_canvas.setFontSize(6)
        my_canvas.setTitle(book_name)
        count = 0
        for serial_number in serial_numbers:
            qr_code = serial_number
            qr_label = qr_code + " " + book_name

            file_path = "/tmp/" + get_random_string(LENGTH_QR_CODE_SVG_FILENAME) + ".svg"
            qr0 = pyqrcode.create(qr_code)
            svg = qr0.svg(file_path)
            write_to_pdf(my_canvas, file_path, qr_label, count)
            import os
            if os.path.exists(file_path):
                os.remove(file_path)
            count += 1
            if count % 63 == 0:
                my_canvas.showPage()
                my_canvas.setFontSize(6)

        my_canvas.save()

        base64_excel_data = base64.b64encode(buffer.getvalue())
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + BOOK_WORKSHEET_NAME + '.xlsx'
        response.write(base64_excel_data)
        return response

    @action(detail=False, methods=['POST'], permission_classes=[AllowAny])
    def alternate_generate_qr_code(self, request, pk=None):
        return Response(status=404)
        is_valid, error_message = validate_user_file_import(request)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        file_in_memory = request.FILES['file'].read()
        workbook = load_workbook(filename=BytesIO(file_in_memory))
        worksheet = workbook[BOOK_WORKSHEET_NAME]
        rows = list(worksheet.rows)
        is_valid, error_message = validate_book_and_inventory_file_excel_content(rows)
        if not is_valid:
            return Response(status=400, data=create_response_data(error_message))

        book_levels = BookLevel.objects.all()
        response = []
        error_in_file = False
        buffer = BytesIO()
        my_canvas = canvas.Canvas(buffer, bottomup=0)
        my_canvas.setFontSize(6)
        my_canvas.setTitle("PMC")

        count = 0

        for index, row in enumerate(rows):
            if index == 0:
                continue
            key = row[0].value.strip() if row[0].value else None
            book_name = row[1].value.strip() if row[1].value else None
            book_level_name = row[2].value.strip() if row[2].value else None

            if book_name is None and key is None and book_level_name is None:
                continue

            book_level = get_book_level(book_levels, book_level_name)
            if not book_level and book_level_name:
                response.append(create_file_upload_error(index, 'level', 'Value must be one of ' + str(
                    get_valid_book_levels())))
                error_in_file = True
                continue
            author = row[3].value.strip() if row[3].value else None
            publisher = row[4].value.strip() if row[4].value else None
            price = str(row[5].value) if row[5].value else None
            serial = str(row[6].value).strip() if row[6].value else None
            copies = int(str(row[7].value).strip()) if row[7].value else None

            # Create inventory
            if not serial:
                response.append(create_file_upload_error(index, 'serial', 'blank serial'))
                error_in_file = True

            if not copies:
                response.append(create_file_upload_error(index, 'copies', 'blank copies'))
                error_in_file = True

            if not copies or copies < 1:
                response.append(create_file_upload_error(index, 'copies', 'invalid copies'))
                error_in_file = True

            # print(serial)
            # print(serial.find(".1", -2))

            #  FOR PMC
            if serial.find(".1", -2) == -1:
                response.append(create_file_upload_error(index, 'serial', '.1 is missing serial'))
                error_in_file = True
                continue

            #  FOR PCMC
            # if serial.find("-0001", -5) == -1:
            #     response.append(create_file_upload_error(index, 'serial', '-0001 is missing serial'))
            #     error_in_file = True
            #     continue

            # FOR PMC

            serial_base = serial[:-1]
            print(index)
            for x in range(1, copies + 1):
                serial_number = serial_base + str(x)
                qr_code = serial_number
                qr_label = qr_code + " " + book_name

                file_path = "/tmp/" + get_random_string(LENGTH_QR_CODE_SVG_FILENAME) + ".svg"
                qr0 = pyqrcode.create(qr_code)
                svg = qr0.svg(file_path)
                write_to_pdf(my_canvas, file_path, qr_label, count)
                import os
                if os.path.exists(file_path):
                    os.remove(file_path)
                count += 1
                if count % 63 == 0:
                    my_canvas.showPage()
                    my_canvas.setFontSize(6)

            # FOR PCMC

            # my_canvas.setTitle(book_name)
            # serial_base = serial[:-5]
            # print(index)
            # for x in range(1, copies + 1):
            #     serial_number = serial_base + '-' + ("%04d" % x)
            #     qr_code = serial_number
            #     qr_label = qr_code + " " + book_name
            #
            #     file_path = "/tmp/" + get_random_string(LENGTH_QR_CODE_SVG_FILENAME) + ".svg"
            #     qr0 = pyqrcode.create(qr_code)
            #     svg = qr0.svg(file_path)
            #     write_to_pdf(my_canvas, file_path, qr_label, count)
            #     import os
            #     if os.path.exists(file_path):
            #         os.remove(file_path)
            #     count += 1
            #     if count % 63 == 0:
            #         my_canvas.showPage()
            #         my_canvas.setFontSize(6)

        if error_in_file:
            raise DatabaseError

        my_canvas.save()

        base64_excel_data = base64.b64encode(buffer.getvalue())
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + BOOK_WORKSHEET_NAME + '.xlsx'
        response.write(base64_excel_data)
        return response


class InventoryViewSet(ViewSet):

    def update(self, request, pk=None):
        if not has_permission(request, PERMISSION_CAN_CHANGE_INVENTORY):
            return Response(status=403, data=ERROR_403_JSON())

        book_key = request.data.get('book')
        try:
            inventory = Inventory.objects.get(key=pk)
            book = Book.objects.get(key=book_key)
        except (Inventory.DoesNotExist, Book.DoesNotExist) as e:
            return Response(status=404, data=create_response_error(e))

        if not request_user_belongs_to_book_ngo(request, book):
            return Response(status=403, data=ERROR_403_JSON())

        data = request.data.copy()
        data["book_id"] = book.id
        serializer = InventorySerializer(inventory, data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(status=400, data=create_serializer_error(serializer))

    @action(methods=['POST'], detail=False, permission_classes=[CanDeleteInventory])
    def deactivate_inventory_books(self, request, pk=None):
        if not request.user.ngo:
            return Response(status=403, data=ERROR_403_JSON())

        # TODO SEND BOOK key too.
        inventory_keys = json.loads(request.data.get("keys"))
        inventory_books = Inventory.objects.filter(key__in=inventory_keys)
        for inventory in inventory_books:
            inventory.is_active = False
            inventory.save()
        return Response(status=204)


class BookLevelViewSet(ViewSet):

    @action(methods=['GET'], detail=False)
    def get_book_levels(self, request):
        levels = convert_to_dropdown(BookLevel.BOOK_LEVELS)
        return Response(levels)
